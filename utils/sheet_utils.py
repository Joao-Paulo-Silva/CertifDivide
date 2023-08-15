import os
import openpyxl
import ezodf

def extract_column_by_header(sheet_file, header):
    _, ext = os.path.splitext(sheet_file)
    
    if ext == '.ods':
        doc = ezodf.opendoc(sheet_file)
        ods_sheet = doc.sheets[0]
        column_index = None
        
        for idx, cell in enumerate(ods_sheet[0]):
            if cell.value == header:
                column_index = idx
                break
        
        if column_index is not None:
            column_values = [row[column_index].value for row in ods_sheet[1:]]
            return column_values
    elif ext == '.xlsx':
        wb = openpyxl.load_workbook(sheet_file)
        xlsx_sheet = wb.active
        column_index = None
        
        for column in xlsx_sheet.iter_cols(min_row=1, max_row=1, values_only=True):
            for idx, value in enumerate(column):
                if value == header:
                    column_index = idx + 1
                    break
        
        if column_index is not None:
            column_values = [row[column_index - 1].value for row in xlsx_sheet.iter_rows(min_row=2, values_only=True)]
            return column_values
    
    return []
